import time
import allure
import pytest
from allure_commons.types import AttachmentType
from appium import webdriver
from appium.webdriver.appium_service import AppiumService
import openpyxl as xl


def get_data():
    workbook = xl.load_workbook("./data_driven/appium_data.xlsx")
    sheet = workbook["Sheet1"]
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    main_list = []
    for i in range(2, total_rows + 1):
        data_list = []
        for j in range(1, total_cols + 1):
            data = sheet.cell(row=i, column=j).value
            data_list.insert(j, data)
        main_list.insert(i, data_list)
    return main_list


@pytest.fixture(params=["device1", "device2"], scope="function")
def appium_driver(request):
    global appium_service
    appium_service = AppiumService()
    appium_service.start()
    if request.params == "device1":
        desired_cap = dict(
            deviceName='Galaxy Note 10',
            platformName='Android',
            platformVersion='11',
            appPackage='com.goibibo',
            udid='R58M86QW34P',
            appActivity='.common.HomeActivity',
            automationName='UiAutomator2',
            noReset=True
        )
    if request.params == "device2":
        desired_cap = dict(
            deviceName='Galaxy Note 10',
            platformName='Android',
            platformVersion='11',
            appPackage='com.goibibo',
            udid='emulator-5554',
            appActivity='.common.HomeActivity',
            automationName='UiAutomator2',
            noReset=True
        )

    global driver
    driver = webdriver.Remote('http://127.0.0.1:4723/wd/hub', desired_cap)
    driver.implicitly_wait(10)
    yield
    driver.quit()
    appium_service.stop()


@pytest.mark.parametrize("city, country", get_data())
def test_goibibo(city, country):
    driver.find_element_by_xpath(
        "//android.widget.LinearLayout[@content-desc='hotels']/android.widget.TextView").click()
    allure.attach(driver.get_screenshot_as_png(), name="screenshot1", attachment_type=AttachmentType.PNG)
    driver.find_element_by_xpath(
        "//android.view.ViewGroup[@content-desc='destination']/android.widget.TextView").click()
    allure.attach(driver.get_screenshot_as_png(), name="screenshot2", attachment_type=AttachmentType.PNG)
    driver.find_element_by_id("com.goibibo:id/edtSearch").send_keys(city)
    allure.attach(driver.get_screenshot_as_png(), name="screenshot3", attachment_type=AttachmentType.PNG)
    dropdown_list = driver.find_elements_by_id("com.goibibo:id/lytLocationItem")
    dropdown_list[0].click()
    driver.find_element_by_xpath(
        "//android.view.ViewGroup[@content-desc='getsetgo_clicked']/android.view.ViewGroup/android.widget.TextView").click()
    allure.attach(driver.get_screenshot_as_png(), name="screenshot4", attachment_type=AttachmentType.PNG)
    time.sleep(3)
    cityText = driver.find_element_by_xpath("//android.widget.TextView[contains(@text,'EXPLORE')]").text
    print(cityText)
    newCityText = str(cityText).replace("EXPLORE ", "").replace("!", "")
    print(newCityText)
    assert newCityText in str(city).upper()
    allure.attach(driver.get_screenshot_as_png(), name="screenshot", attachment_type=AttachmentType.PNG)
