import openpyxl as xl
from openpyxl import workbook

xl.load_workbook("data_driven/test_excel.xlsx")
sheet = workbook["Sheet1"]
total_row = sheet.max_row
total_column = sheet.max_column
cell_value = sheet(row=1, column=1).value
